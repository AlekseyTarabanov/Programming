from math import sqrt
import openpyxl
from openpyxl import Workbook
import numpy as np
import matplotlib.pyplot as plt
import task2 # хочу взять 1 функцию из 2 звдвния(вспомнить как реализовывать модули)
import random

def random_cor(count_cor, diopozon_x, diopozon_y):
	cor_list = [] 
	global x_list, y_list
	for i in range(1, count_cor+1):
		x = 'A' + str(i+1)
		y = 'B' + str(i+1)
		while ws[x].value == None:
			a = random.randint(diopozon_x[0], diopozon_x[1])
			b = random.randint(diopozon_y[0], diopozon_y[1])
			if [a,b] in cor_list:
				continue
			else:
				ws[x] = str(a)
				ws[y] = str(b)
				x_list.append(a)
				y_list.append(b)
				cor_list.append([a,b])
	print("-"*20,"\n Метод наименьших квадратов\n",MNK(x_list, y_list))
				
def MNK(x_list, y_list):
	x_mid = sum(x_list)/len(x_list)
	y_mid = sum(y_list)/len(y_list)
	x_otklon = 0
	num = 0
	for i in range(len(x_list)):
		x_otklon = x_otklon + (x_list[i] - x_mid)
		num += (x_list[i] - x_mid)*(y_list[i] - y_mid)
	b1, b2 = 0, 0

	if x_otklon != 0:
		b2 = num/(x_otklon**2)
		b1 = y_mid - (b2*x_mid)
		return "Yi = " + str(b1) + "+" + str(b2) + "* Xi"
	else:
		return "Yi = Xi"
					
if __name__ == '__main__':
	wb = Workbook()
	ws = wb.active
	ws['A1'] = 'X'
	ws['B1'] = 'Y'
	count_cor = int(sqrt(2) * 10) # 2 вариант
	x_list, y_list = [],[]
	diopozon_x_input = list(input('Введите диапазон координат X в порядке возрастания, через запятую: '))
	diopozon_y_input = list(input('Введите диапазон координат Y в порядке возрастания, через запятую: '))
	diopozon_x, diopozon_y = [], []
	diopozon_x = task2.perevod(diopozon_x_input, diopozon_x)
	diopozon_y = task2.perevod(diopozon_y_input, diopozon_y)
	for i in range(0,len(diopozon_x)):
		diopozon_x[i] = int(diopozon_x[i])
	for i in range(0,len(diopozon_y)):
		diopozon_y[i] = int(diopozon_y[i])
		
	if ((diopozon_x[1] - diopozon_x[0]) * (diopozon_y[1] - diopozon_y[0])) <= 14:
		print("Вы ввели слишком маленький диапазон координат")
	else:
		random_cor(count_cor, diopozon_x, diopozon_y)
		print("-"*20, "\nВсё готово, смотри таблицу")
		wb.save("For Task5.xls")
		plt.plot(x_list, y_list)
		plt.show()
